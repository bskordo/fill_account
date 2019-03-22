import openpyxl
import argparse
import os
import sys
import json
import requests
import time


filepath = os.path.join(sys.path[0], "purch_work.json")


def load_json_data(filepath):
    if not os.path.exists(filepath):
        return None
    with open(filepath, 'r') as file_handler:
        return json.load(file_handler)


def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('file_which_need_to_fill', help='file_which_need_to_fill')
    arg = parser.parse_args()
    return arg.file_which_need_to_fill


def get_stop_fact_from_dadata(inn):
    headers = {'Content-type': 'application/json',
    'Accept' : 'application/json','Authorization': 'Token d9d064f66d6a3ac3702461be3133cd62f13158d8'}
    data = '{"query": "%(inn)s"}'% {"inn": inn}
    response = requests.post('https://suggestions.dadata.ru/suggestions/api/4_1/rs/findById/party',
     headers=headers, data=data)
    if response.status_code == 200:
        my_response =response.json()
        try:
            data_response = my_response['suggestions'][0]
            fias_address=[
            data_response['data']['state']
            ]
            return fias_address[0]
        except:
            return None
    else:
        return None

def gos_contracts(inn):
    s=inn
    headers = {'Authorization': 'Bearer HWqOvOrVMHAWS7N6MBli3tRxW7xlPzEglk5nZDgi'}
    response = requests.get('https://agent.smartbank.pro/api/gos/contracts/?supplier_inn=%s'%s,headers=headers)
    if response.status_code == 200:
        contracts = response.json()['data']
        return contracts
    else:
        return None

def get_okved_from_dadata(code):
    headers = {'Content-type': 'application/json',
    'Accept' : 'application/json','Authorization': 'Token d9d064f66d6a3ac3702461be3133cd62f13158d8'}
    data = '{"query": "%(code)s"}'% {"code": code}
    response = requests.post('https://suggestions.dadata.ru/suggestions/api/4_1/rs/findById/okved2',
     headers=headers, data=data)
    if response.status_code == 200:
        my_response =response.json()
        try:
            data_response = my_response['suggestions'][0]
            fias_address=[
            data_response['data']['name']
            ]
            return fias_address[0]
        except:
            return None
    else:
        return None
 

def write_infromation_into_file(my_file, inf_fr_jsn,stop_inf_dadata, gos_contracts):
    work_book = openpyxl.load_workbook(filename=my_file)
    #work_sheet = work_book.active
    work_sheet = work_book.worksheets[0]
    work_sheet['B4'] = inf_fr_jsn['principal']['full_name']
    work_sheet['D18'] = inf_fr_jsn['principal']['short_name']
    work_sheet['D19'] = inf_fr_jsn['principal']['inn']
    work_sheet['D20'] = inf_fr_jsn['principal']['kpp']
    work_sheet['D21'] = inf_fr_jsn['principal']['ogrn']
    work_sheet['F23'] = inf_fr_jsn['principal']['creation_date']
    work_sheet['F24'] = inf_fr_jsn['principal']['registration_place']
    #Адрес местонахождения указанный при
    work_sheet['F26'] = inf_fr_jsn['principal']['legal_address']['data']['postal_code']
    work_sheet['F27'] = inf_fr_jsn['principal']['legal_address']['data']['region']
    work_sheet['F28'] = inf_fr_jsn['principal']['legal_address']['data']['city_district']
    work_sheet['F29'] = inf_fr_jsn['principal']['legal_address']['data']['city_with_type']
    work_sheet['F30'] = inf_fr_jsn['principal']['legal_address']['data']['settlement_with_type']
    work_sheet['F31'] = inf_fr_jsn['principal']['legal_address']['data']['street']
    work_sheet['F32'] = inf_fr_jsn['principal']['legal_address']['data']['house']
    work_sheet['F33'] = inf_fr_jsn['principal']['legal_address']['data']['block']
    if inf_fr_jsn['principal']['legal_address']['data']['block'] =='оф':
        work_sheet['F35'] = inf_fr_jsn['principal']['legal_address']['data']['flat']


    #End
    #Контакты Компании
    work_sheet['F37'] = inf_fr_jsn['principal']['email']
    work_sheet['F38'] = inf_fr_jsn['principal']['contact_phone']
    #СВЕДЕНИЯ ОБ ОТВЕТСТВЕННЫХ ЛИЦАХ
    work_sheet['F40'] = inf_fr_jsn['principal']['management_name']
    if inf_fr_jsn['principal']['CEO']['is_name_changed'] is True:
        work_sheet['F41'] = 'ДА'
    else:
        work_sheet['F41'] = 'НЕТ'
    work_sheet['F42'] = inf_fr_jsn['principal']['CEO']['citizenship']
    work_sheet['F43'] = inf_fr_jsn['principal']['CEO']['birth_place']
    work_sheet['F44'] = inf_fr_jsn['principal']['CEO']['birth_date']
    work_sheet['F45'] = inf_fr_jsn['principal']['CEO']['snils']
    work_sheet['F46'] = inf_fr_jsn['principal']['CEO']['inn']
    work_sheet['D48'] = inf_fr_jsn['principal']['okved_main']
    okved_main = get_okved_from_dadata(inf_fr_jsn['principal']['okved_main'])
    work_sheet['D49'] = okved_main
    work_sheet['D51'] = inf_fr_jsn['principal']['employees_number']
    #3.0. Сведения о заявке 
    work_sheet['D64'] = 'ПОТ Экспресс - гарантия'
    work_sheet['D65'] = inf_fr_jsn['guarantee_amount']
    work_sheet['D66'] = 'БГ по '+inf_fr_jsn['purchase_law']+' Ф3'
    if inf_fr_jsn['has_prepayment'] is True:
        work_sheet['D67'] = 'ДА'
    else:
        work_sheet['D67'] = 'НЕТ'
    if inf_fr_jsn['is_big_deal'] is True:
        work_sheet['D68'] = 'ДА'
    else:
        work_sheet['D68'] = 'НЕТ'
    work_sheet['D64'] = 'ПОТ Экспресс - гарантия'

    work_sheet['D69'] = 'Исполнение контракта'
    work_sheet['G64'] = inf_fr_jsn['guarantee_type_label']
    work_sheet['G65'] = inf_fr_jsn['guarantee_start_date']
    work_sheet['G66'] = inf_fr_jsn['guarantee_end_date']
    guarantee_amount = float(inf_fr_jsn['guarantee_amount'])
    purchase_price = float(inf_fr_jsn['purchase_starting_price'])
    percent = (guarantee_amount/purchase_price)*100
    fin_percent = str(round(percent,1))+'%'
    work_sheet['G67'] = fin_percent
    work_sheet['G69'] = 'Типовая'
    work_sheet['D73'] = inf_fr_jsn['beneficiary']['full_name']
    work_sheet['D74'] = inf_fr_jsn['beneficiary']['short_name']
    work_sheet['D75'] = inf_fr_jsn['beneficiary']['inn']
    work_sheet['D76'] = inf_fr_jsn['beneficiary']['kpp']
    work_sheet['D77'] = inf_fr_jsn['beneficiary']['ogrn']
    work_sheet['D79'] = inf_fr_jsn['beneficiary']['legal_address']['value']
    work_sheet['D81'] = inf_fr_jsn['purchase_number']
    work_sheet['D82'] = inf_fr_jsn['purchase_subject']
    work_sheet['D83'] = inf_fr_jsn['purchase_url']
    work_sheet['D85'] = inf_fr_jsn['purchase_starting_price']
    work_sheet['D86'] = '-'
    work_sheet['D87'] = '-'
    work_sheet['E88'] = '-'
    work_sheet['E89'] = '-'
    work_sheet['G88'] = '-'
    work_sheet['G89'] = '-'
    #5.0. Проверка Принципала на безусловные параметры СТОП-информации 
    #Недействительность паспорта единоличного исполнительного органа
    if inf_fr_jsn['principal']['guarantee_scoring']['stop_factors'][4]['pass']:
        work_sheet['G94'] = 'да'
    else:
        work_sheet['G94'] = 'нет'
    #Отсутствие регистрации в ЕГРЮЛ
    if inf_fr_jsn['principal']['inn']:
        work_sheet['G95'] = 'нет'
    else:
        work_sheet['G95'] = 'да'
    #Сведения о проведении процедуры ликвидации или банкротства в отношении Принципала
    if stop_inf_dadata['liquidation_date']:
        work_sheet['G96'] = 'да'
    else:
        work_sheet['G96'] = 'нет'
    #Принципал - недействующее юридическое лицо:
    if stop_inf_dadata['status']=='ACTIVE':
        work_sheet['G97'] = 'нет'
    else:
        work_sheet['G97'] = 'да'
    #Дисквалификация исполнительного органа
    work_sheet['G98'] = 'нет' #TODO need to clarify

    #Недобросовестность Поставщика (принципал находится в реестре недобросовестных поставщиков)
    if inf_fr_jsn['principal']['guarantee_scoring']['stop_factors'][2]['pass']:
        work_sheet['G99'] = 'да'
    else:
        work_sheet['G99'] = 'нет'
    #Наличие исполнительных производств по кредитным договорам
    work_sheet['G100'] = 'нет' #TODO need to clarify
    #Наличие прочих исполнительных производств, превышающих установленные значения
    work_sheet['G101'] = 'нет' #TODO need to clarify
    #Наличие арбитражных дел в отношении Принципала (Принципал выступает в качестве ответчика)
    if inf_fr_jsn['principal']['guarantee_scoring']['stop_factors'][17]['pass']:
        work_sheet['G102'] = 'да'
    else:
        work_sheet['G102'] = 'нет'
    #Наличие просроченной задолженности по налогам и сборам
    work_sheet['G103'] = 'нет'#TODO need to clarify
    #Значится в перечне террористов / экстремистов
    if inf_fr_jsn['principal']['guarantee_scoring']['stop_factors'][8]['pass']:
        work_sheet['G104'] = 'да'
    else:
        work_sheet['G104'] = 'нет'
    #Значится в перечне неблагонадежных участников ВЭД в соответствии с Письмом Банка России 193-Т
    if inf_fr_jsn['principal']['guarantee_scoring']['stop_factors'][6]['pass']:
        work_sheet['G105'] = 'да'
    else:
        work_sheet['G105'] = 'нет'
    #Отрицательная кредитная история
    work_sheet['G106'] = 'нет'#TODO need to clarify
    #Сумма гарантии превышает лимит на Принципала
    work_sheet['G107'] = 'нет'#TODO need to clarify
    #Наличие у компании отрицательных чистых активов и текущего и/или накопленного убытка на последнюю отчетную дату
    first_stg=inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][9]['values']['30.09.2018']
    sec_stg=inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][7]['values']['30.09.2018']
    third_stg=inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][4]['values']['30.09.2018']
    fourth_stg=inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][2]['fields'][6]['values']['30.09.2018']
    assets = float(first_stg)+float(sec_stg)-float(third_stg)-float(fourth_stg)
    if assets > 0:
        work_sheet['G108'] = 'нет'
    else:
        work_sheet['G108'] = 'да'
    #Наличие у компании текущего убытка и накопленного убытка на последнюю отчетную дату
    lost=inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][11]['values']['30.09.2018']
    if float(lost) > 0:
        work_sheet['G109'] = 'нет'
    else:
        work_sheet['G109'] = 'да'
    #ЕИО сменился более двух раз за последний год
    work_sheet['G110'] = 'нет'#TODO need to clarify
    #Место постановки Принципала на налоговый учет изменилось более двух раз
    work_sheet['G111'] = 'нет'



    #EndПроверка Принципала на безусловные параметры СТОП-информации
    if gos_contracts:
        contracts = work_book.worksheets[9]
        for contract in range(1,len(gos_contracts)):
            contracts.cell(row=contract+1, column=1).value = gos_contracts[contract-1]['law']
            contracts.cell(row=contract+1, column=2).value = gos_contracts[contract-1]['purchaseNumber']
            contracts.cell(row=contract+1, column=3).value = gos_contracts[contract-1]['regNumber']
            contracts.cell(row=contract+1, column=4).value = gos_contracts[contract-1]['lotNumber']
            contracts.cell(row=contract+1, column=5).value = gos_contracts[contract-1]['purchaseObject']
            contracts.cell(row=contract+1, column=6).value = gos_contracts[contract-1]['supplierName']
            signDate = gos_contracts[contract-1]['signDate']['$date']
            sign_date = time.strftime("%d %m %Y", time.gmtime(signDate / 1000.0))
            contracts.cell(row=contract+1, column=7).value = sign_date
            contracts.cell(row=contract+1, column=8).value = gos_contracts[contract-1]['price']
            StartDate = gos_contracts[contract-1]['executionStartDate']['$date']
            start_date = time.strftime("%d %m %Y", time.gmtime(StartDate / 1000.0))
            contracts.cell(row=contract+1, column=9).value = start_date
            EndDate = gos_contracts[contract-1]['executionEndDate']['$date']
            end_date = time.strftime("%d %m %Y", time.gmtime(EndDate / 1000.0))
            contracts.cell(row=contract+1, column=11).value = end_date
            contracts.cell(row=contract+1, column=12).value = gos_contracts[contract-1]['stage']
              
    #БУХ ОТЧЕТ
    count_sheet = work_book.worksheets[1]
    count_sheet['B2'] = inf_fr_jsn['principal']['full_name']
    #Нематериальные активы_1110
    count_sheet['D8'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][0]['values']['31.12.2017']
    count_sheet['E8'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][0]['values']['31.03.2018']
    count_sheet['F8'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][0]['values']['30.06.2018']
    count_sheet['G8'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][0]['values']['30.09.2018']

    #Результаты исследований и разработок_1120
    count_sheet['D9'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][1]['values']['31.12.2017']
    count_sheet['E9'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][1]['values']['31.03.2018']
    count_sheet['F9'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][1]['values']['30.06.2018']
    count_sheet['G9'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][1]['values']['30.09.2018']

    #Нематериальные поисковые активы_1130
    count_sheet['D10'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][2]['values']['31.12.2017']
    count_sheet['E10'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][2]['values']['31.03.2018']
    count_sheet['F10'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][2]['values']['30.06.2018']
    count_sheet['G10'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][0]['fields'][2]['values']['30.09.2018']

    #Запасы_1210
    count_sheet['D13'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][0]['values']['31.12.2017']
    count_sheet['E13'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][0]['values']['31.03.2018']
    count_sheet['F13'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][0]['values']['30.06.2018']
    count_sheet['G13'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][0]['values']['30.09.2018']
    
    #НДС по приобретенным ценностям 1220
    count_sheet['D14'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][1]['values']['31.12.2017']
    count_sheet['E14'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][1]['values']['31.03.2018']
    count_sheet['F14'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][1]['values']['30.06.2018']
    count_sheet['G14'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][0]['groups'][1]['fields'][1]['values']['30.09.2018']

    #Резервный капитал_1360
    count_sheet['D17'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][4]['values']['31.12.2017']
    count_sheet['E17'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][4]['values']['31.03.2018']
    count_sheet['F17'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][4]['values']['30.06.2018']
    count_sheet['G17'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][4]['values']['30.09.2018']

    #Нераспределенная прибыль (непокрытый убыток)_1370
    count_sheet['D18'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][5]['values']['31.12.2017']
    count_sheet['E18'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][5]['values']['31.03.2018']
    count_sheet['F18'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][5]['values']['30.06.2018']
    count_sheet['G18'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][0]['fields'][5]['values']['30.09.2018']

    #Заёмные средства   1410
    count_sheet['D21'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][0]['values']['31.12.2017']
    count_sheet['E21'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][0]['values']['31.03.2018']
    count_sheet['F21'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][0]['values']['30.06.2018']
    count_sheet['G21'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][0]['values']['30.09.2018']
    
    #Отложенные налоговые обязательства_1420
    count_sheet['D22'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][1]['values']['31.12.2017']
    count_sheet['E22'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][1]['values']['31.03.2018']
    count_sheet['F22'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][1]['values']['30.06.2018']
    count_sheet['G22'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][1]['values']['30.09.2018']

    #Оценочные обязательства_1430
    count_sheet['D23'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][2]['values']['31.12.2017']
    count_sheet['E23'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][2]['values']['31.03.2018']
    count_sheet['F23'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][2]['values']['30.06.2018']
    count_sheet['G23'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][2]['values']['30.09.2018']

    #Прочие обязательства_1450
    count_sheet['D26'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][3]['values']['31.12.2017']
    count_sheet['E26'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][3]['values']['31.03.2018']
    count_sheet['F26'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][3]['values']['30.06.2018']
    count_sheet['G26'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][1]['groups'][1]['fields'][3]['values']['30.09.2018']

    #Выручка
    count_sheet['D32'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][0]['values']['31.12.2017']
    count_sheet['E32'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][0]['values']['31.03.2018']
    count_sheet['F32'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][0]['values']['30.06.2018']
    count_sheet['G32'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][0]['values']['30.09.2018']

    #Себестоимость продаж
    count_sheet['D33'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][1]['values']['31.12.2017']
    count_sheet['E33'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][1]['values']['31.03.2018']
    count_sheet['F33'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][1]['values']['30.06.2018']
    count_sheet['G33'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][0]['fields'][1]['values']['30.09.2018']

     #Доходы от участия в других организациях_2310
    count_sheet['D42'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][0]['values']['31.12.2017']
    count_sheet['E42'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][0]['values']['31.03.2018']
    count_sheet['F42'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][0]['values']['30.06.2018']
    count_sheet['G42'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][0]['values']['30.09.2018']
    
    #Проценты к получению_2320
    count_sheet['D43'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][1]['values']['31.12.2017']
    count_sheet['E43'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][1]['values']['31.03.2018']
    count_sheet['F43'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][1]['values']['30.06.2018']
    count_sheet['G43'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][1]['values']['30.09.2018']
    
    #Проценты к уплате_2330
    count_sheet['D44'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][2]['values']['31.12.2017']
    count_sheet['E44'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][2]['values']['31.03.2018']
    count_sheet['F44'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][2]['values']['30.06.2018']
    count_sheet['G44'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][2]['values']['30.09.2018']

    #Прочие доходы_2340
    count_sheet['D46'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][3]['values']['31.12.2017']
    count_sheet['E46'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][3]['values']['31.03.2018']
    count_sheet['F46'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][3]['values']['30.06.2018']
    count_sheet['G46'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][3]['values']['30.09.2018']

    #Прочие доходы_2350
    count_sheet['D47'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][4]['values']['31.12.2017']
    count_sheet['E47'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][4]['values']['31.03.2018']
    count_sheet['F47'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][4]['values']['30.06.2018']
    count_sheet['G47'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][4]['values']['30.09.2018']

    #Текущий налог на прибыль_2410
    count_sheet['D50'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][6]['values']['31.12.2017']
    count_sheet['E50'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][6]['values']['31.03.2018']
    count_sheet['F50'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][6]['values']['30.06.2018']
    count_sheet['G50'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][6]['values']['30.09.2018']

    #Изменение отложенных налоговых обязательств_2430
    count_sheet['D52'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][8]['values']['31.12.2017']
    count_sheet['E52'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][8]['values']['31.03.2018']
    count_sheet['F52'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][8]['values']['30.06.2018']
    count_sheet['G52'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][8]['values']['30.09.2018']

    #Прочее_2460
    count_sheet['D55'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][10]['values']['31.12.2017']
    count_sheet['E55'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][10]['values']['31.03.2018']
    count_sheet['F55'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][10]['values']['30.06.2018']
    count_sheet['G55'] = inf_fr_jsn['principal']['accountability']['data']['fieldsets'][2]['groups'][1]['fields'][10]['values']['30.09.2018']


    #КОНЕЦ БУХ ОТЧЕТА



    work_book.save(my_file)


if __name__ == '__main__':
    file_name = get_args()
    inf_fr_jsn = load_json_data(filepath)
    stop_inf_dadata = get_stop_fact_from_dadata(inf_fr_jsn['principal']['inn'])
    gos_contracts = gos_contracts(inf_fr_jsn['principal']['inn'])
    write_infromation_into_file(file_name, inf_fr_jsn,stop_inf_dadata,gos_contracts)